import os
import re
import sys
import openpyxl
import xml.etree.ElementTree as ET

def capitalizar_fuente(fuente):
    if not fuente:
        return ""
    # Regla 11: Capitalización adecuada de nombres y organizaciones
    # Omitimos preposiciones en minúsculas en medio de las frases
    lower_exceptions = ['de', 'la', 'el', 'los', 'las', 'del', 'en', 'y', 'con', 'por', 'para', 'a']
    words = fuente.split()
    capitalized_words = []
    for idx, w in enumerate(words):
        clean_w = re.sub(r'[^\wáéíóúñÁÉÍÓÚÑ]', '', w, flags=re.UNICODE)
        if clean_w.lower() in lower_exceptions and idx > 0:
            capitalized_words.append(w.lower())
        else:
            if len(w) > 0:
                first_alpha_match = re.search(r'[a-zA-ZáéíóúñÁÉÍÓÚÑ]', w)
                if first_alpha_match:
                    start_idx = first_alpha_match.start()
                    capitalized_words.append(w[:start_idx] + w[start_idx].upper() + w[start_idx+1:].lower())
                else:
                    capitalized_words.append(w.lower())
            else:
                capitalized_words.append(w)
    return " ".join(capitalized_words)

def generar_glosario_xml(excel_path, xml_output_path):
    print(f"Cargando archivo Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    
    # Buscar la pestaña del glosario
    sheet_name = None
    for name in wb.sheetnames:
        if 'glosario' in name.lower():
            sheet_name = name
            break
            
    if not sheet_name:
        print("Error: No se encontró ninguna pestaña que contenga 'glosario' en el libro de Excel.")
        return
        
    sheet = wb[sheet_name]
    print(f"Procesando hoja: {sheet.title}")
    
    # Buscar dinámicamente dónde inician las filas de datos
    start_row = 8  # Fallback
    for r in range(1, 25):
        val = sheet.cell(row=r, column=2).value
        if val and 'verbo' in str(val).lower():
            start_row = r + 1
            break
            
    entries = []
    current_dimension = "General"
    
    for r_idx in range(start_row, sheet.max_row + 1):
        dim_val = sheet.cell(row=r_idx, column=1).value
        verb_val = sheet.cell(row=r_idx, column=2).value
        acepcion_val = sheet.cell(row=r_idx, column=3).value
        fuente_val = sheet.cell(row=r_idx, column=5).value
        
        if dim_val:
            current_dimension = str(dim_val).strip().capitalize()
            
        if verb_val:
            verb = str(verb_val).strip()
            verb_concept = verb.capitalize()
            acepcion = str(acepcion_val).strip() if acepcion_val else ""
            fuente = str(fuente_val).strip() if fuente_val else ""
            
            # Formatear RAE y capitalizar fuentes
            fuente_capitalized = capitalizar_fuente(fuente)
            if "https://dle.rae.es" in fuente_capitalized.lower():
                raw_url = f"https://dle.rae.es/{verb.lower()}"
                anchor = f'<a href="{raw_url}" target="_blank" rel="noreferrer noopener">{raw_url}</a>'
                fuente_formatted = re.sub(r'<?https://dle\.rae\.es>?', anchor, fuente_capitalized, flags=re.IGNORECASE)
            else:
                fuente_formatted = fuente_capitalized
                
            # Generar tabla HTML limpia de Bootstrap
            table_html = f'''<table class="table table-striped table-bordered">
    <tbody>
        <tr>
            <th style="text-align:center;">DIMENSIÓN</th>
            <th style="text-align:center;">VERBO</th>
            <th style="text-align:center;">ACEPCIÓN</th>
            <th style="text-align:center;">FUENTE</th>
        </tr>
        <tr>
            <td>{current_dimension}</td>
            <td>{verb_concept}</td>
            <td>{acepcion}</td>
            <td>{fuente_formatted}</td>
        </tr>
    </tbody>
</table>
<br>
<p></p>'''
            entries.append({
                'concept': verb_concept,
                'definition': table_html,
                'category': current_dimension.upper()
            })
            
    # Ordenar alfabéticamente
    entries.sort(key=lambda x: x['concept'].lower())
    
    # Generar estructura XML con CDATA y tags estándar de Moodle
    xml_lines = []
    xml_lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    xml_lines.append('<GLOSSARY>')
    xml_lines.append('  <INFO>')
    xml_lines.append('    <NAME>Glosario de Verbos del Curso</NAME>')
    xml_lines.append('    <INTRO>Importación automática del Glosario de Verbos de Rúbricas.</INTRO>')
    xml_lines.append('  </INFO>')
    
    # Definición de categorías a nivel raíz
    xml_lines.append('  <CATEGORIES>')
    for cat_name in ["SABER", "SER", "HACER"]:
        xml_lines.append('    <CATEGORY>')
        xml_lines.append(f'      <NAME>{cat_name}</NAME>')
        xml_lines.append('      <USEDYNALINK>0</USEDYNALINK>')
        xml_lines.append('    </CATEGORY>')
    xml_lines.append('  </CATEGORIES>')
    
    xml_lines.append('  <ENTRIES>')
    
    for entry in entries:
        xml_lines.append('    <ENTRY>')
        xml_lines.append(f'      <CONCEPT>{entry["concept"]}</CONCEPT>')
        xml_lines.append('      <DEFINITION><![CDATA[')
        xml_lines.append(entry["definition"])
        xml_lines.append('      ]]></DEFINITION>')
        xml_lines.append('      <FORMAT>1</FORMAT>')
        xml_lines.append('      <USEDYNALINK>0</USEDYNALINK>')
        xml_lines.append('      <CASESENSITIVE>0</CASESENSITIVE>')
        xml_lines.append('      <FULLMATCH>0</FULLMATCH>')
        xml_lines.append('      <TEACHERENTRY>1</TEACHERENTRY>')
        xml_lines.append('      <CATEGORIES>')
        xml_lines.append(f'        <CATEGORY>{entry["category"]}</CATEGORY>')
        xml_lines.append('      </CATEGORIES>')
        xml_lines.append('    </ENTRY>')
        
    xml_lines.append('  </ENTRIES>')
    xml_lines.append('</GLOSSARY>')
    
    with open(xml_output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(xml_lines))
        
    print(f"\nSe ha generado con éxito el archivo XML con CDATA: {xml_output_path}")
    print(f"Se crearon {len(entries)} términos listos para importar en Moodle.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python generar_glosario_xml.py [ruta_al_excel_rubrica.xlsx]")
        sys.exit(1)
        
    excel = sys.argv[1]
    output = os.path.abspath(os.path.join(os.path.dirname(excel), "../../3_paginas_finales/glosario_importar.xml"))
    generar_glosario_xml(excel, output)
